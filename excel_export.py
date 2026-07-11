"""
Excel export of the equipment schedule, styled to match the IAE branding
(Inglese Architecture + Engineering FHA report formatting template):

  - Header row: maroon fill (#5F3844), white bold text
  - Alternating body rows: white / light pink (#F5EDEF)
  - Font: Swis721 LtCn BT (falls back to Calibri if not installed on the
    opening machine — Swis721 is not a standard Windows/Mac font, so ship
    the .xlsx with a documented fallback rather than assume it's present)
  - Thin borders on every cell, color #333333
  - Column widths auto-sized to content
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IAE_MAROON = "5F3844"
IAE_WHITE = "FFFFFF"
IAE_BODY = "222222"
IAE_BORDER = "333333"
IAE_ALT_ROW = "F5EDEF"
FONT_NAME = "Swis721 LtCn BT"  # falls back automatically if not installed


def build_schedule_workbook(rows: list, columns: list, title: str = "Mechanical Equipment Schedule") -> bytes:
    """
    rows: list of dicts, each {column_key: value}
    columns: the discipline's column schema (see schema.py)
    title: schedule title, written as a banner row above the table

    Returns the .xlsx file as bytes, ready to hand to Streamlit's download_button
    or to open directly and link into AutoCAD as an OLE / data link object.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    thin = Side(style="thin", color=IAE_BORDER)
    border = Border(top=thin, left=thin, bottom=thin, right=thin)

    n_cols = len(columns)

    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=title.upper())
    title_cell.font = Font(name=FONT_NAME, size=13, bold=True, color=IAE_WHITE)
    title_cell.fill = PatternFill("solid", fgColor=IAE_MAROON)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Header row
    header_row_idx = 2
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=col["label"])
        cell.font = Font(name=FONT_NAME, size=12, bold=True, color=IAE_WHITE)
        cell.fill = PatternFill("solid", fgColor=IAE_MAROON)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[header_row_idx].height = 30

    # Body rows
    for r_offset, row_data in enumerate(rows):
        row_idx = header_row_idx + 1 + r_offset
        is_alt = (r_offset % 2 == 1)
        fill_color = IAE_ALT_ROW if is_alt else IAE_WHITE
        for col_idx, col in enumerate(columns, start=1):
            value = row_data.get(col["key"], "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name=FONT_NAME, size=12, color=IAE_BODY)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    # Column widths — auto-size based on label length and a sane minimum
    for col_idx, col in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        header_len = len(col["label"])
        max_data_len = max(
            [len(str(row_data.get(col["key"], ""))) for row_data in rows] + [header_len]
        )
        ws.column_dimensions[letter].width = max(12, min(max_data_len + 4, 30))

    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
